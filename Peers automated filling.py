import yahoo_fin.stock_info as si 
from openpyxl import *
import pandas as pd
from bs4 import BeautifulSoup
import requests, re, time

##yahoo package
##http://theautomatic.net/2020/05/05/how-to-download-fundamentals-data-with-python/
##pd.set_option('display.max_columns', None)

def get_dividend_yield(symbol):
    data = si.get_quote_table(symbol , dict_result = True)
    begin = data['Forward Dividend & Yield'].index('(')
    try:
        end = data['Forward Dividend & Yield'].index('%')    
    except:
        return "N/A"
    return (float(data['Forward Dividend & Yield'][begin+1:end]))/100

def Float(symbol):
    url = requests.get('https://finance.yahoo.com/quote/'+symbol+'/key-statistics?p='+symbol).text
    soup = BeautifulSoup(url,'lxml')
    alldata = soup.find_all('tbody')
    try:
        table = alldata[2].find_all('tr')
    except:
        table = None
        
    result = str(table[3].text)
    
    return result.split(' ')[1]

def Company_name(symbol):
    url = "http://d.yimg.com/autoc.finance.yahoo.com/autoc?query={}&region=1&lang=en".format(symbol)

    result = requests.get(url).json()

    for x in result['ResultSet']['Result']:
        if x['symbol'] == symbol:
            return x['name']    


column_A = ['Múltiplos',
            'Sales (ttm)',
            'Sales (last 10-K)',
            'Growth (ttm)',
            'Growth (36m)',
            'Debt',
            'Asset',
            'Debt/Asset',
            'Debt/Ebitda',
            '(Capex + Adq.)/Revenue',
            'Operating Margin (TTM)',
            'Float',
            'Beta (5Y Monthly)',
            'PE Ratio (TTM)',
            'Price/Book (mrq)',
            'EV/Asset',
            'EV/Ebitda',
            'Dividend Yield (%)']
col = 2

workbook = Workbook()
sheet = workbook.active

##column 1
for i in range(len(column_A)):
    sheet.cell(row = i+1, column = 1).value = column_A[i]
        
print('Type Q to quit the program')
print('')
print('')

while True:
    symbol = input('Write the Ticker Symbol of the company: ')
    symbol = symbol.upper()
    
    if symbol == 'Q':
        break

    if Company_name(symbol) == None:
        print('Company not found')
        print('')
        col += 1
        continue

    print('Company: ' + Company_name(symbol))

    print('Collecting Data: ', end = '')
    start_time = time.time()
    Balance_Sheet = si.get_balance_sheet(symbol)
    Income_Statement = si.get_income_statement(symbol)
    Cash_Flow = si.get_cash_flow(symbol)
    Quote = si.get_quote_table(symbol)
    Valuation_Stats = si.get_stats_valuation(symbol)
    print(round(time.time() - start_time,2), 'seconds')
    
    ##column 2
    print('Insert Data: ', end = '')
    start_time = time.time()
    total_assets = int(Balance_Sheet.loc[list(Balance_Sheet.iloc[:,0]).index('Total Assets')][1])
    total_debt = int(Balance_Sheet.loc[list(Balance_Sheet.iloc[:,0]).index('Total Debt')][1])
    
    try: ebitda_anual = int(Income_Statement.loc[list(Income_Statement.iloc[:,0]).index('Normalized EBITDA')][2])
    except: ebitda_anual = 'Fail'

    if ebitda_anual == 'Fail': Debt_Ebitda = 'Fail'
    else: Debt_Ebitda = total_debt / ebitda_anual
    
    try: operating_income = int(Income_Statement.loc[list(Income_Statement.iloc[:,0]).index('Operating Income')][1])
    except: operating_income = 'Fail'

    if operating_income == 'Fail': operating_margin = 'Fail'
    else: operating_margin = operating_income / int(Income_Statement.loc[0][1])
    
    PB_ratio = float(Valuation_Stats.loc[list(Valuation_Stats.iloc[:,0]).index('Price/Book (mrq)')][1])

    for i in range(len(Valuation_Stats.columns)):
        if (Valuation_Stats.columns[i] == Balance_Sheet.columns[1]) or (Balance_Sheet.columns[1] in Valuation_Stats.columns[i]):
            EV_ebitda_anual_col_index = i

    EV_ebitda_anual = float(Valuation_Stats.loc[list(Valuation_Stats.iloc[:,0]).index('Enterprise Value/EBITDA 6')][EV_ebitda_anual_col_index])

    if ebitda_anual == 'Fail': EV_Assets = 'Fail'
    else: EV_Assets = (EV_ebitda_anual * ebitda_anual) / total_assets

    sheet.cell(row = 1, column = col).value = symbol
    sheet.cell(row = 2, column = col).value = Income_Statement.loc[0][1]
    sheet.cell(row = 3, column = col).value = Income_Statement.loc[0][2]
    sheet.cell(row = 4, column = col).value = (float(sheet.cell(row = 2, column = col).value) / float(sheet.cell(row = 3, column = col).value)) - 1
    sheet.cell(row = 4, column = col).number_format = '0.00%'
    sheet.cell(row = 5, column = col).value = (float(sheet.cell(row = 3, column = col).value) / float(Income_Statement.loc[0][4])) - 1
    sheet.cell(row = 5, column = col).number_format = '0.00%'
    sheet.cell(row = 6, column = col).value = total_debt
    sheet.cell(row = 7, column = col).value = total_assets
    sheet.cell(row = 8, column = col).value = total_debt / total_assets
    sheet.cell(row = 8, column = col).number_format = '0.00'
    sheet.cell(row = 9, column = col).value = Debt_Ebitda
    sheet.cell(row = 9, column = col).number_format = '0.00'
    sheet.cell(row = 11, column = col).value = operating_margin
    sheet.cell(row = 11, column = col).number_format = '0.00%'
    sheet.cell(row = 12, column = col).value = Float(symbol)
    sheet.cell(row = 13, column = col).value = float(Quote['Beta (5Y Monthly)'])
    sheet.cell(row = 14, column = col).value = Quote["PE Ratio (TTM)"]
    sheet.cell(row = 15, column = col).value = PB_ratio
    sheet.cell(row = 16, column = col).value = EV_Assets
    sheet.cell(row = 16, column = col).number_format = '0.00'
    sheet.cell(row = 17, column = col).value = EV_ebitda_anual
    sheet.cell(row = 18, column = col).value = get_dividend_yield(symbol)
    sheet.cell(row = 18, column = col).number_format = '0.00%'

    #Verification
    sheet.cell(row = 21, column = col).value = Company_name(symbol)
    
    if (Balance_Sheet.columns[1] != Income_Statement.columns[2]):
        sheet.cell(row = 22, column = col).value = 'Fail'
    else:
        sheet.cell(row = 22, column = col).value = 'OK'

    if (Valuation_Stats.columns[EV_ebitda_anual_col_index] == Income_Statement.columns[2]) and (Valuation_Stats.columns[EV_ebitda_anual_col_index] == Balance_Sheet.columns[1]):
        sheet.cell(row = 23, column = col).value = 'OK'
    else:
        sheet.cell(row = 23, column = col).value = 'Fail'

    sheet.cell(row = 24, column = col).value = Income_Statement.columns.values[1]
    sheet.cell(row = 25, column = col).value = Income_Statement.columns.values[2]
    sheet.cell(row = 26, column = col).value = Income_Statement.columns.values[4]
    
    print(round(time.time() - start_time,2), 'seconds')
    col += 1
    print('')
    print('')

#column 1: verification
sheet.cell(row = 20, column = 1).value = 'Verification'
sheet.cell(row = 21, column = 1).value = 'Company'
sheet.cell(row = 22, column = 1).value = column_A[8]
sheet.cell(row = 23, column = 1).value = column_A[15]
sheet.cell(row = 24, column = 1).value = 'Sales (ttm)'
sheet.cell(row = 25, column = 1).value = 'Sales (last 10-K)'
sheet.cell(row = 26, column = 1).value = 'Sales (36M)'

print('')
file_name = input('What is the name of the file? ')
file_name = str(file_name) + '.xlsx'

print('')
print('Saving File')
workbook.save(filename = file_name)
time.sleep(1)

print('Exiting in 3 seconds')
time.sleep(3)
